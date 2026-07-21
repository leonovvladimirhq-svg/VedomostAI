"""Генерация ведомости в Excel (.xlsx) — контур 1/3, задача «сначала обычный Excel».

Источник истины — наш движок (grading_service). Excel — только представление/выгрузка.
Дальше эта же структура ляжет на Яндекс.Таблицы (тот же макет колонок).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.services.grading_service import Scheme, compute

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_BOLD = Font(bold=True)


def build_ledger(
    scheme: Scheme,
    students: list[str],
    *,
    course_name: str = "",
    module: str = "",
    entries: dict[str, dict[str, list[float]]] | None = None,
) -> Workbook:
    """students — список ФИО. entries[ФИО][element_key] = список оценок (может отсутствовать).
    Возвращает Workbook: №, ФИО, колонка на элемент контроля, Итог."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    title = f"{course_name} — {module}".strip(" —")
    if title:
        ws.cell(1, 1, title).font = Font(bold=True, size=13)

    header = ["№", "ФИО"] + [f"{e.name} (вес {e.weight:g})" for e in scheme.elements] + ["Итог"]
    hrow = 2
    for c, name in enumerate(header, start=1):
        cell = ws.cell(hrow, c, name)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, fio in enumerate(students, start=1):
        row = hrow + i
        ws.cell(row, 1, i)
        ws.cell(row, 2, fio)
        stu_entries = (entries or {}).get(fio, {})
        for ci, el in enumerate(scheme.elements, start=3):
            vals = stu_entries.get(el.key, [])
            ws.cell(row, ci, round(sum(vals) / len(vals), 2) if vals else None)
        res = compute(scheme, stu_entries)
        ws.cell(row, 3 + len(scheme.elements), res.total)

    # ширины
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    for ci in range(3, 3 + len(scheme.elements) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    return wb


def save_ledger(path: str, *args, **kwargs) -> str:
    build_ledger(*args, **kwargs).save(path)
    return path


def build_ledger_from_statement(session, statement) -> Workbook:
    """Генерирует Excel по ведомости из БД: студенты × элементы + Итог (через движок)."""
    from core.models import Group
    from core.services import statement_service as svc

    group = session.get(Group, statement.group_id)
    students = svc.roster(session, group)
    els = svc.scheme_elements(session, statement)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"
    title = f"{statement.course_name} — {statement.module}".strip(" —")
    if title:
        ws.cell(1, 1, title).font = Font(bold=True, size=13)

    header = ["№", "ФИО"] + [f"{e.name} ({e.weight:g})" for e in els] + ["Итог"]
    for c, name in enumerate(header, start=1):
        cell = ws.cell(2, c, name)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, stu in enumerate(students, start=1):
        row = 2 + i
        ws.cell(row, 1, i)
        ws.cell(row, 2, stu.full_name)
        ent = svc.entries_for_student(session, statement, stu)
        res = svc.student_total(session, statement, stu)
        for ci, e in enumerate(els, start=3):
            has = bool(ent.get(str(e.id)))
            ws.cell(row, ci, round(res.aggregated.get(str(e.id), 0), 2) if has else None)
        ws.cell(row, 3 + len(els), res.total)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    for ci in range(3, 3 + len(els) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    return wb
