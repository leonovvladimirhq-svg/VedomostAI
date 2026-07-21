"""Тесты расчётного движка — самый ответственный модуль (план, раздел 5, DoD ≥95%)."""
from __future__ import annotations

import openpyxl
import pytest

from core.parsing.pud_parser import parse_formula
from core.services.grading_service import (
    Element, Scheme, compute, round_half_away, seminar_scheme,
)

XLSX = r"C:/Users/Владимир/Desktop/Ведомость_3-4_модули.xlsx"


# --- Округление (Excel ROUND: половина от нуля) ---
@pytest.mark.parametrize("x,exp", [(4.5, 5), (4.44, 4), (2.5, 3), (4.8, 5), (0.0, 0)])
def test_round_half_away(x, exp):
    assert round_half_away(x, 0) == exp


def test_rounding_modes():
    s_arith = Scheme([Element("a", "A", 1.0, "single")], rounding="arithmetic")
    s_down = Scheme([Element("a", "A", 1.0, "single")], rounding="down")
    s_pass4 = Scheme([Element("a", "A", 1.0, "single")], rounding="arithmetic_pass4")
    assert compute(s_arith, {"a": [3.5]}).total == 4      # арифметическое: вверх
    assert compute(s_down, {"a": [3.9]}).total == 3       # всегда вниз
    assert compute(s_pass4, {"a": [3.5]}).total == 3      # <4 -> вниз (4 проходной)
    assert compute(s_pass4, {"a": [4.5]}).total == 5      # >=4 -> арифметическое


def test_gate_to_zero():
    """Нет посещений -> итог 0 независимо от прочего (как в Excel: V=ЕСЛИ(R=0;0;U))."""
    s = seminar_scheme()
    res = compute(s, {"seminars": [], "report": [10]})
    assert res.total == 0 and res.zeroed_by_gate


def test_blocking_caps_total():
    s = Scheme([
        Element("main", "Основной", 0.5, "single"),
        Element("exam", "Экзамен", 0.5, "single", is_blocking=True),
    ])
    # 0.5*10 + 0.5*6 = 8 -> округл 8, но блокирующий=3 => итог не выше 3
    res = compute(s, {"main": [10], "exam": [3]})
    assert res.total == 3


def test_pud_formula_parsing():
    f = "Активность * 0.1 + Блиц * 0.2 + Тест: Викторина * 0.2 + Контрольная работа * 0.5"
    scheme = parse_formula(f)
    assert len(scheme.elements) == 4
    assert scheme.check_weights() == 1.0
    assert scheme.elements[-1].weight == 0.5


def test_matches_real_excel_number_by_number():
    """Движок воспроизводит реальную Excel-ведомость по всем студентам."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    scheme = seminar_scheme()
    checked = 0
    for r in range(3, ws.max_row + 1):
        if not ws.cell(r, 2).value:
            continue
        sessions = [float(ws.cell(r, c).value) for c in range(4, 18)
                    if isinstance(ws.cell(r, c).value, (int, float))]
        rep = ws.cell(r, 20).value
        report = [float(rep)] if isinstance(rep, (int, float)) else []
        expected = float(ws.cell(r, 22).value)
        assert compute(scheme, {"seminars": sessions, "report": report}).total == expected
        checked += 1
    assert checked >= 40
