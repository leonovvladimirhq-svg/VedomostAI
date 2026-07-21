"""Валидация расчётного движка против реальной Excel-ведомости «число в число».

Читает Ведомость_3-4_модули.xlsx, для каждого студента собирает оценки за занятия
(колонки D:Q) и Отчёт (T), считает нашим движком и сверяет с колонкой Итог (V).
"""
from __future__ import annotations

import sys

import openpyxl

from core.services.grading_service import compute, seminar_scheme

XLSX = r"C:/Users/Владимир/Desktop/Ведомость_3-4_модули.xlsx"


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    scheme = seminar_scheme()

    total, ok = 0, 0
    mism = []
    for r in range(3, ws.max_row + 1):
        surname = ws.cell(r, 2).value
        if not surname:
            continue
        # занятия: колонки D(4)..Q(17)
        sessions = [ws.cell(r, c).value for c in range(4, 18)]
        sessions = [float(v) for v in sessions if isinstance(v, (int, float))]
        # отчёт: колонка T(20)
        report = ws.cell(r, 20).value
        report = [float(report)] if isinstance(report, (int, float)) else []
        expected = ws.cell(r, 22).value  # V = Итог

        res = compute(scheme, {"seminars": sessions, "report": report})
        total += 1
        got = res.total
        if float(expected) == float(got):
            ok += 1
        else:
            mism.append((r, surname, sessions, report, got, expected))

    print(f"Сверено студентов: {total}, совпало: {ok}, расхождений: {len(mism)}")
    for r, s, sess, rep, got, exp in mism[:15]:
        print(f"  строка {r} {s}: занятия={sess} отчёт={rep} -> движок={got} excel={exp}")
    return 0 if ok == total else 1


if __name__ == "__main__":
    sys.exit(main())
